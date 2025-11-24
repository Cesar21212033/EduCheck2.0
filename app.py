import os
import re
import json
import base64
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import atexit
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import qrcode
from qrcode.image.pil import PilImage
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Cargar variables de entorno desde c.env (o .env si existe)
# Primero intentar cargar c.env, luego .env (tiene prioridad si existe)
env_file_c = os.path.join(os.path.dirname(__file__), 'c.env')
env_file_dot = os.path.join(os.path.dirname(__file__), '.env')

if os.path.exists(env_file_c):
    load_dotenv(env_file_c, override=False)  # Cargar c.env primero
if os.path.exists(env_file_dot):
    load_dotenv(env_file_dot, override=True)  # .env tiene prioridad si existe
else:
    # Si no existe .env, intentar cargar c.env de nuevo para asegurar que se carga
    load_dotenv(env_file_c, override=True)

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET", "dev-secret")
# Example: mysql+pymysql://user:password@localhost/asistencia_db
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv("DATABASE_URL", "mysql+pymysql://root:Baby20150531@localhost/asistencia_db")
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Configuración de correo
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER', 'smtp.gmail.com')  # Gmail como servidor por defecto
app.config['MAIL_PORT'] = int(os.getenv('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.getenv('MAIL_USE_TLS', 'True').lower() in ['true', '1', 'yes']
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME', 'cesarini.05ramos@gmail.com')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD', '').strip()  # Eliminar espacios en blanco
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_DEFAULT_SENDER', 'cesarini.05ramos@gmail.com')
# Configuración adicional para Flask-Mail
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_SUPPRESS_SEND'] = False  # Permitir envío real de correos

db = SQLAlchemy(app)
mail = Mail(app)

# ---------------------
# MODELOS
# ---------------------
class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    nombre_completo = db.Column(db.String(150), nullable=False)
    rol = db.Column(db.String(20), nullable=False, default='profesor')  # admin, profesor, alumno
    activo = db.Column(db.Boolean, default=True)
    password_changed = db.Column(db.Boolean, default=False)  # Para alumnos: indica si ya cambió su contraseña inicial
    creado_en = db.Column(db.DateTime, default=datetime.utcnow)
    # Relación con clases (materias)
    clases = db.relationship('Clase', backref='maestro', lazy=True, cascade='all, delete-orphan')
    # Relación con estudiante (si es alumno)
    estudiante = db.relationship('Estudiante', backref='usuario', uselist=False, cascade='all, delete-orphan')

    def set_password(self, password):
        """Genera hash de la contraseña"""
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        """Verifica la contraseña"""
        return check_password_hash(self.password_hash, password)
    
    def is_admin(self):
        """Verifica si el usuario es administrador"""
        return self.rol == 'admin'
    
    def is_profesor(self):
        """Verifica si el usuario es profesor"""
        return self.rol == 'profesor'
    
    def is_alumno(self):
        """Verifica si el usuario es alumno"""
        return self.rol == 'alumno'
    
    @property
    def es_admin(self):
        """Propiedad para usar en templates Jinja2"""
        return self.rol == 'admin'
    
    @property
    def es_profesor(self):
        """Propiedad para usar en templates Jinja2"""
        return self.rol == 'profesor'
    
    @property
    def es_alumno(self):
        """Propiedad para usar en templates Jinja2"""
        return self.rol == 'alumno'

class Estudiante(db.Model):
    __tablename__ = 'estudiantes'
    id = db.Column(db.Integer, primary_key=True)
    numero_control = db.Column(db.String(50), unique=True, nullable=False)
    nombre = db.Column(db.String(150), nullable=False)
    correo = db.Column(db.String(150))
    qr_code = db.Column(db.Text, nullable=True)  # Almacenar QR como base64
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id', ondelete='SET NULL'), nullable=True)
    creado_en = db.Column(db.DateTime, default=datetime.utcnow)
    # Relación muchos a muchos con clases
    clases = db.relationship('Clase', secondary='estudiante_clase', back_populates='estudiantes', lazy='dynamic')

class Clase(db.Model):
    __tablename__ = 'clases'
    id = db.Column(db.Integer, primary_key=True)
    codigo_clase = db.Column(db.String(50), unique=True, nullable=False)
    nombre_clase = db.Column(db.String(150), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    creado_en = db.Column(db.DateTime, default=datetime.utcnow)
    # Relación muchos a muchos con estudiantes
    estudiantes = db.relationship('Estudiante', secondary='estudiante_clase', back_populates='clases', lazy='dynamic')
    # Relación uno a muchos con horarios
    horarios = db.relationship('HorarioClase', backref='clase', lazy='dynamic', cascade='all, delete-orphan')

class HorarioClase(db.Model):
    __tablename__ = 'horarios_clase'
    id = db.Column(db.Integer, primary_key=True)
    clase_id = db.Column(db.Integer, db.ForeignKey('clases.id', ondelete='CASCADE'), nullable=False)
    dia_semana = db.Column(db.Integer, nullable=False)  # 1=Lunes, 2=Martes, 3=Miércoles, 4=Jueves, 5=Viernes
    hora_inicio = db.Column(db.Time, nullable=False)
    hora_fin = db.Column(db.Time, nullable=False)
    creado_en = db.Column(db.DateTime, default=datetime.utcnow)
    
    __table_args__ = (
        db.UniqueConstraint('clase_id', 'dia_semana', name='unique_clase_dia'),
    )

class ReporteEnviado(db.Model):
    __tablename__ = 'reportes_enviados'
    id = db.Column(db.Integer, primary_key=True)
    clase_id = db.Column(db.Integer, db.ForeignKey('clases.id', ondelete='CASCADE'), nullable=False)
    fecha_clase = db.Column(db.Date, nullable=False)
    enviado_en = db.Column(db.DateTime, default=datetime.utcnow)
    
    __table_args__ = (
        db.UniqueConstraint('clase_id', 'fecha_clase', name='unique_reporte_dia'),
    )

# Tabla de asociación muchos a muchos entre Estudiante y Clase
estudiante_clase = db.Table('estudiante_clase',
    db.Column('estudiante_id', db.Integer, db.ForeignKey('estudiantes.id', ondelete='CASCADE'), primary_key=True),
    db.Column('clase_id', db.Integer, db.ForeignKey('clases.id', ondelete='CASCADE'), primary_key=True),
    db.Column('asignado_en', db.DateTime, default=datetime.utcnow)
)

class Asistencia(db.Model):
    __tablename__ = 'asistencias'
    id = db.Column(db.Integer, primary_key=True)
    estudiante_id = db.Column(db.Integer, db.ForeignKey('estudiantes.id'), nullable=False)
    clase_id = db.Column(db.Integer, db.ForeignKey('clases.id'), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    hora = db.Column(db.Time, nullable=False)
    metodo = db.Column(db.String(50), default='QR')
    info_extra = db.Column(db.JSON, nullable=True)
    creado_en = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relaciones
    estudiante = db.relationship('Estudiante', backref='asistencias')
    clase = db.relationship('Clase', backref='asistencias')

    __table_args__ = (db.UniqueConstraint('estudiante_id', 'clase_id', 'fecha', name='unica_asistencia_por_dia'),)

# ---------------------
# FUNCIONES AUXILIARES
# ---------------------

def validar_email(email):
    """Valida formato de email, acepta correos @tectijuana.edu.mx y otros dominios válidos"""
    if not email:
        return True  # Email es opcional en algunos contextos
    
    email = email.strip()
    
    # Validación muy simple y permisiva
    # Patrón básico: algo@algo.algo
    # Acepta prácticamente cualquier formato de email válido
    
    # Verificaciones básicas
    if not email or len(email) < 5:  # Mínimo: a@b.c
        return False
    
    if ' ' in email:  # No espacios
        return False
    
    if email.count('@') != 1:  # Exactamente un @
        return False
    
    # Separar usuario y dominio
    try:
        usuario, dominio = email.split('@', 1)
    except ValueError:
        return False
    
    # Validar usuario: no vacío, puede tener letras, números, puntos, guiones, etc.
    if not usuario or len(usuario) == 0:
        return False
    
    # Validar dominio: debe tener al menos un punto
    if not dominio or '.' not in dominio:
        return False
    
    # Verificar que el dominio termine con al menos 2 letras (TLD)
    partes_dominio = dominio.split('.')
    if len(partes_dominio) < 2:
        return False
    
    tld = partes_dominio[-1]
    if len(tld) < 2:
        return False
    
    # Si pasa todas las verificaciones básicas, es válido
    return True

def validar_numero_control(numero):
    """Valida formato de número de control"""
    if not numero:
        return False
    # Permitir alfanuméricos, guiones y guiones bajos, entre 3 y 50 caracteres
    pattern = r'^[a-zA-Z0-9_-]{3,50}$'
    return re.match(pattern, numero) is not None

def validar_nombre(nombre):
    """Valida formato de nombre"""
    if not nombre:
        return False
    # Permitir letras, espacios, acentos, entre 2 y 150 caracteres
    pattern = r'^[a-zA-ZáéíóúÁÉÍÓÚñÑüÜ\s]{2,150}$'
    return re.match(pattern, nombre) is not None

def validar_codigo_clase(codigo):
    """Valida formato de código de clase"""
    if not codigo:
        return False
    # Permitir alfanuméricos, guiones y guiones bajos, entre 2 y 50 caracteres
    pattern = r'^[a-zA-Z0-9_-]{2,50}$'
    return re.match(pattern, codigo) is not None

def validar_username(username):
    """Valida formato de username"""
    if not username:
        return False
    # Permitir alfanuméricos y guiones bajos, entre 3 y 30 caracteres
    pattern = r'^[a-zA-Z0-9_]{3,30}$'
    return re.match(pattern, username) is not None

# ---------------------
# DECORADORES Y FUNCIONES DE AUTENTICACIÓN
# ---------------------

def login_required(f):
    """Decorador para requerir autenticación"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor, inicia sesión para acceder a esta página', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def get_current_user():
    """Obtiene el usuario actual desde la sesión"""
    if 'user_id' in session:
        return Usuario.query.get(session['user_id'])
    return None

def admin_required(f):
    """Decorador para requerir rol de administrador"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor, inicia sesión para acceder a esta página', 'warning')
            return redirect(url_for('login'))
        usuario = get_current_user()
        if not usuario or not usuario.is_admin():
            flash('No tienes permisos para acceder a esta página', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def profesor_required(f):
    """Decorador para requerir rol de profesor"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor, inicia sesión para acceder a esta página', 'warning')
            return redirect(url_for('login'))
        usuario = get_current_user()
        if not usuario or (not usuario.is_profesor() and not usuario.is_admin()):
            flash('No tienes permisos para acceder a esta página', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def alumno_required(f):
    """Decorador para requerir rol de alumno"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor, inicia sesión para acceder a esta página', 'warning')
            return redirect(url_for('login'))
        usuario = get_current_user()
        if not usuario or not usuario.is_alumno():
            flash('No tienes permisos para acceder a esta página', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def generar_qr_estudiante(numero_control):
    """Genera QR para un número de control y lo retorna como base64"""
    payload = {
        "numero_control": numero_control
    }
    payload_str = json.dumps(payload)
    
    # Generar QR usando Pillow directamente
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(payload_str)
    qr.make(fit=True)
    
    # Usar image_factory para asegurar que use Pillow (PIL)
    img = qr.make_image(fill_color="black", back_color="white", image_factory=PilImage)
    
    buf = BytesIO()
    img.save(buf, 'PNG')  # Usar formato como segundo argumento posicional
    buf.seek(0)
    
    # Convertir a base64 para almacenar en BD
    qr_base64 = base64.b64encode(buf.read()).decode('utf-8')
    return qr_base64

def enviar_correo_bienvenida(usuario, password_temporal):
    """Envía correo de bienvenida a admin o profesor"""
    if not usuario.email:
        return False, "El usuario no tiene correo electrónico registrado"
    
    # Verificar configuración de correo
    mail_password = app.config.get('MAIL_PASSWORD', '') or os.getenv('MAIL_PASSWORD', '')
    mail_username = app.config.get('MAIL_USERNAME', '') or os.getenv('MAIL_USERNAME', '')
    
    if not mail_password or mail_password.strip() == '':
        app.logger.warning('MAIL_PASSWORD no configurada, no se puede enviar correo de bienvenida')
        return False, "Configuración de correo incompleta"
    
    try:
        año_actual = datetime.now().year
        rol_nombre = 'Administrador' if usuario.rol == 'admin' else 'Profesor'
        
        msg = Message(
            subject=f'Bienvenido a EduCheck - {rol_nombre}',
            recipients=[usuario.email],
            html=render_template('email_bienvenida.html', 
                               nombre_completo=usuario.nombre_completo,
                               username=usuario.username,
                               email=usuario.email,
                               rol=usuario.rol,
                               rol_nombre=rol_nombre,
                               password_temporal=password_temporal,
                               año_actual=año_actual),
            sender=app.config.get('MAIL_DEFAULT_SENDER', mail_username)
        )
        
        mail.init_app(app)
        mail.send(msg)
        
        return True, "Correo de bienvenida enviado exitosamente"
    except Exception as e:
        error_msg = str(e)
        app.logger.error(f'Error al enviar correo de bienvenida: {error_msg}')
        return False, f"Error al enviar correo: {error_msg}"

def enviar_qr_por_correo(estudiante, qr_base64):
    """Envía el QR del estudiante por correo electrónico"""
    if not estudiante.correo:
        return False, "El estudiante no tiene correo electrónico registrado"
    
    # Verificar que la configuración de correo esté completa
    # Intentar obtener desde app.config primero, luego desde os.getenv directamente
    mail_password = app.config.get('MAIL_PASSWORD', '') or os.getenv('MAIL_PASSWORD', '')
    mail_username = app.config.get('MAIL_USERNAME', '') or os.getenv('MAIL_USERNAME', '')
    
    # Log para depuración (sin mostrar la contraseña completa)
    app.logger.info(f'Verificando configuración de correo - Username: {mail_username}, Password configurada: {"Sí" if mail_password else "No"}')
    
    if not mail_password or mail_password.strip() == '':
        app.logger.warning('Configuración de correo incompleta: MAIL_PASSWORD no está configurada')
        return False, "Configuración de correo incompleta. Por favor configure MAIL_PASSWORD en el archivo c.env (o .env) con la contraseña de su cuenta de correo."
    
    if not mail_username or mail_username.strip() == '':
        app.logger.warning('Configuración de correo incompleta: MAIL_USERNAME no está configurada')
        return False, "Configuración de correo incompleta. Verifique MAIL_USERNAME en el archivo c.env (o .env)"
    
    try:
        # Decodificar el QR desde base64
        qr_bytes = base64.b64decode(qr_base64)
        
        # Crear el mensaje dentro del contexto de la aplicación
        año_actual = datetime.now().year
        msg = Message(
            subject=f'Código QR de Asistencia - {estudiante.nombre}',
            recipients=[estudiante.correo],
            html=render_template('email_qr.html', estudiante=estudiante, año_actual=año_actual),
            sender=app.config.get('MAIL_DEFAULT_SENDER', mail_username)
        )
        
        # Adjuntar el QR como imagen
        msg.attach(
            filename=f'qr_{estudiante.numero_control}.png',
            content_type='image/png',
            data=qr_bytes
        )
        
        # Enviar el correo
        # Flask-Mail maneja el contexto automáticamente cuando se llama desde una ruta Flask
        # Asegurar que las credenciales estén actualizadas en la configuración de mail
        mail.init_app(app)
        mail.send(msg)
        
        return True, "QR enviado exitosamente"
    except Exception as e:
        error_msg = str(e)
        error_lower = error_msg.lower()
        app.logger.error(f'Error al enviar correo: {error_msg}')
        app.logger.error(f'Configuración actual - Server: {app.config.get("MAIL_SERVER")}, Port: {app.config.get("MAIL_PORT")}, Username: {app.config.get("MAIL_USERNAME")}')
        
        # Mensajes de error más específicos
        if 'connect' in error_lower or 'connection' in error_lower or 'timeout' in error_lower:
            return False, "Error de conexión con el servidor de correo. Verifique MAIL_SERVER, MAIL_PORT y su conexión a internet."
        elif 'authentication' in error_lower or 'login' in error_lower or '535' in error_msg or '535-5.7.3' in error_msg:
            mail_server = app.config.get('MAIL_SERVER', 'smtp.gmail.com')
            if 'gmail' in mail_server.lower():
                return False, "Error de autenticación con Gmail. Gmail requiere una 'Contraseña de aplicación' en lugar de tu contraseña normal. Ve a https://myaccount.google.com/apppasswords y genera una contraseña de aplicación. Úsala en MAIL_PASSWORD."
            else:
                return False, "Error de autenticación. Necesitas generar una 'Contraseña de aplicación' desde la configuración de seguridad de tu proveedor de correo y usarla en MAIL_PASSWORD."
        elif 'password' in error_lower or 'credentials' in error_lower:
            mail_server = app.config.get('MAIL_SERVER', 'smtp.gmail.com')
            if 'gmail' in mail_server.lower():
                return False, "Credenciales incorrectas. Para Gmail, necesitas generar una 'Contraseña de aplicación' desde https://myaccount.google.com/apppasswords y usarla en MAIL_PASSWORD."
            else:
                return False, "Credenciales incorrectas. Verifica que estés usando una contraseña de aplicación, no tu contraseña normal."
        elif '535' in error_msg:
            mail_server = app.config.get('MAIL_SERVER', 'smtp.gmail.com')
            if 'gmail' in mail_server.lower():
                return False, "Error 535: Autenticación fallida. Gmail requiere una contraseña de aplicación. Genera una en: https://myaccount.google.com/apppasswords"
            else:
                return False, "Error 535: Autenticación fallida. Verifica que estés usando una contraseña de aplicación."
        else:
            return False, f"Error al enviar correo: {error_msg}. Si usas Gmail, asegúrate de usar una contraseña de aplicación desde https://myaccount.google.com/apppasswords"

# ---------------------
# RUTAS DE AUTENTICACIÓN
# ---------------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('Usuario y contraseña son obligatorios', 'danger')
            return redirect(url_for('login'))
        
        # Buscar usuario por username primero
        usuario = Usuario.query.filter_by(username=username, activo=True).first()
        
        # Si no se encuentra por username, intentar buscar por número de control (para alumnos)
        if not usuario:
            estudiante = Estudiante.query.filter_by(numero_control=username.upper()).first()
            if estudiante and estudiante.usuario_id:
                usuario = Usuario.query.filter_by(id=estudiante.usuario_id, activo=True).first()
        
        if not usuario:
            flash('Usuario o contraseña incorrectos', 'danger')
            return redirect(url_for('login'))
        
        # Verificar que el usuario tenga un password_hash válido
        if not usuario.password_hash or len(usuario.password_hash.strip()) == 0:
            app.logger.error(f'Usuario {username} tiene password_hash vacío o inválido')
            flash('Error: Usuario con contraseña inválida. Contacte al administrador.', 'danger')
            return redirect(url_for('login'))
        
        try:
            if usuario.check_password(password):
                session['user_id'] = usuario.id
                session['username'] = usuario.username
                session['nombre'] = usuario.nombre_completo
                session['rol'] = usuario.rol
                flash(f'Bienvenido, {usuario.nombre_completo}', 'success')
                
                # Si el usuario no ha cambiado su contraseña, redirigir a cambiar contraseña
                if not usuario.password_changed:
                    flash('Por favor, cambia tu contraseña antes de continuar', 'info')
                    return redirect(url_for('cambiar_contraseña'))
                
                # Redirigir según el rol
                if usuario.is_alumno():
                    return redirect(url_for('alumno_dashboard'))
                else:
                    return redirect(url_for('dashboard'))
            else:
                flash('Usuario o contraseña incorrectos', 'danger')
                return redirect(url_for('login'))
        except ValueError as e:
            app.logger.error(f'Error al verificar contraseña para usuario {username}: {str(e)}')
            flash('Error: Contraseña inválida en la base de datos. Contacte al administrador.', 'danger')
            return redirect(url_for('login'))
        except Exception as e:
            app.logger.error(f'Error inesperado en login para usuario {username}: {str(e)}')
            flash('Error al iniciar sesión. Intente nuevamente.', 'danger')
            return redirect(url_for('login'))
    
    # Si ya está logueado, redirigir al dashboard
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    return render_template('login.html')

# La ruta de registro público ha sido eliminada
# Solo los administradores pueden crear usuarios:
# - Profesores: /admin/profesores/nuevo (solo admin)
# - Alumnos: se crean automáticamente al registrar un estudiante en /estudiantes/nuevo

@app.route('/logout')
def logout():
    session.clear()
    flash('Sesión cerrada exitosamente', 'info')
    return redirect(url_for('login'))

# ---------------------
# RUTAS PRINCIPALES
# ---------------------

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    usuario = get_current_user()
    
    # Redirigir alumnos a su dashboard específico
    if usuario.is_alumno():
        return redirect(url_for('alumno_dashboard'))
    
    # Dashboard para admin y profesor
    if usuario.is_admin():
        # Admin ve todas las clases
        clases = Clase.query.order_by(Clase.nombre_clase).all()
        total_estudiantes = Estudiante.query.count()
        total_asistencias_hoy = Asistencia.query.filter_by(fecha=date.today()).count()
    else:
        # Profesor ve solo sus clases
        clases = Clase.query.filter_by(usuario_id=usuario.id).order_by(Clase.nombre_clase).all()
        total_estudiantes = Estudiante.query.count()
        total_asistencias_hoy = Asistencia.query.filter_by(fecha=date.today()).count()
    
    return render_template('dashboard.html', 
                         usuario=usuario, 
                         clases=clases,
                         total_estudiantes=total_estudiantes,
                         total_asistencias_hoy=total_asistencias_hoy)

# Dashboard para alumnos
@app.route('/alumno/dashboard')
@alumno_required
def alumno_dashboard():
    usuario = get_current_user()
    estudiante = Estudiante.query.filter_by(usuario_id=usuario.id).first()
    
    if not estudiante:
        flash('No se encontró información del estudiante', 'danger')
        return redirect(url_for('logout'))
    
    # Obtener clases del estudiante
    clases_estudiante = estudiante.clases.all()
    
    # Obtener asistencias del estudiante
    asistencias = Asistencia.query.filter_by(estudiante_id=estudiante.id).order_by(Asistencia.fecha.desc(), Asistencia.hora.desc()).limit(10).all()
    
    return render_template('alumno_dashboard.html',
                         usuario=usuario,
                         estudiante=estudiante,
                         clases=clases_estudiante,
                         asistencias=asistencias)

# Registro público de estudiantes (solo con correo @sujv.mx)
@app.route('/registro-estudiante', methods=['GET', 'POST'])
def registro_estudiante():
    # Si ya está autenticado, redirigir al dashboard
    if 'user_id' in session:
        usuario = get_current_user()
        if usuario:
            if usuario.is_alumno():
                return redirect(url_for('alumno_dashboard'))
            else:
                return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        num = request.form.get('numero_control', '').strip().upper()
        nombre = request.form.get('nombre', '').strip()
        correo = request.form.get('correo', '').strip().lower()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validaciones
        errores = []
        
        # Validar número de control
        if not num:
            errores.append('El número de control es obligatorio')
        elif not validar_numero_control(num):
            errores.append('El número de control debe tener entre 3 y 50 caracteres alfanuméricos (puede incluir guiones y guiones bajos)')
        
        # Validar nombre
        if not nombre:
            errores.append('El nombre es obligatorio')
        elif not validar_nombre(nombre):
            errores.append('El nombre debe contener solo letras y espacios, entre 2 y 150 caracteres')
        
        # Validar correo - DEBE terminar en @sujv.mx, @gmail.com, @hotmail.com o @tectijuana.edu.mx
        dominios_permitidos = ['@sujv.mx', '@gmail.com', '@hotmail.com', '@tectijuana.edu.mx']
        if not correo:
            errores.append('El correo electrónico es obligatorio')
        elif not validar_email(correo):
            errores.append('El formato del correo electrónico no es válido')
        elif not any(correo.endswith(dominio) for dominio in dominios_permitidos):
            errores.append('Solo se permiten correos electrónicos que terminen en @sujv.mx, @gmail.com, @hotmail.com o @tectijuana.edu.mx')
        
        # Validar contraseña
        if not password:
            errores.append('La contraseña es obligatoria')
        elif len(password) < 6:
            errores.append('La contraseña debe tener al menos 6 caracteres')
        
        if password != confirm_password:
            errores.append('Las contraseñas no coinciden')
        
        # Verificar si ya existe un estudiante con ese número de control (case-insensitive)
        existe_estudiante = Estudiante.query.filter(
            db.func.upper(Estudiante.numero_control) == num.upper()
        ).first()
        
        if existe_estudiante:
            errores.append(f'Ya existe un estudiante con el número de control "{existe_estudiante.numero_control}"')
        
        # Verificar si ya existe un usuario con ese correo (case-insensitive)
        if correo:
            existe_usuario_correo = Usuario.query.filter(
                db.func.lower(Usuario.email) == correo.lower()
            ).first()
            
            if existe_usuario_correo:
                errores.append(f'Ya existe un usuario registrado con el correo electrónico "{correo}". No se pueden registrar correos duplicados.')
        
        # Verificar si ya existe un usuario con ese username (número de control)
        existe_username = Usuario.query.filter_by(username=num.upper()).first()
        if existe_username:
            errores.append(f'Ya existe un usuario con el número de control "{num}"')
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            return render_template('registro_estudiante.html', 
                                 numero_control=num, 
                                 nombre=nombre, 
                                 correo=correo)
        
        # Generar QR
        try:
            qr_base64 = generar_qr_estudiante(num)
        except Exception as e:
            flash(f'Error al generar QR: {str(e)}', 'danger')
            return render_template('registro_estudiante.html',
                                 numero_control=num,
                                 nombre=nombre,
                                 correo=correo)
        
        # Crear usuario tipo alumno
        username_alumno = num.upper()
        # Asegurar que el username sea único
        contador = 1
        username_original = username_alumno
        while Usuario.query.filter_by(username=username_alumno).first():
            username_alumno = f"{username_original}_{contador}"
            contador += 1
        
        # Crear usuario alumno
        usuario_alumno = Usuario(
            username=username_alumno,
            email=correo,
            nombre_completo=nombre,
            rol='alumno',
            password_changed=True  # Ya estableció su contraseña durante el registro
        )
        usuario_alumno.set_password(password)
        db.session.add(usuario_alumno)
        db.session.flush()
        
        # Crear estudiante vinculado al usuario
        est = Estudiante(
            numero_control=num,
            nombre=nombre,
            correo=correo,
            qr_code=qr_base64,
            usuario_id=usuario_alumno.id
        )
        db.session.add(est)
        db.session.commit()
        
        # Enviar QR por correo
        if correo:
            exito, mensaje = enviar_qr_por_correo(est, qr_base64)
            if exito:
                flash(f'Registro exitoso. Tu código QR ha sido enviado a {correo}. Puedes iniciar sesión ahora.', 'success')
            else:
                flash(f'Registro exitoso, pero hubo un problema al enviar el correo: {mensaje}. Tu código QR está disponible en tu panel.', 'warning')
        else:
            flash('Registro exitoso con QR generado', 'success')
        
        return redirect(url_for('login'))
    
    return render_template('registro_estudiante.html')

# Registrar estudiante (solo admin)
@app.route('/estudiantes/nuevo', methods=['GET','POST'])
@admin_required
def nuevo_estudiante():
    if request.method == 'POST':
        num = request.form.get('numero_control', '').strip().upper()  # Convertir a mayúsculas para consistencia
        nombre = request.form.get('nombre', '').strip()
        correo = request.form.get('correo', '').strip()
        
        # Validaciones
        errores = []
        
        if not num:
            errores.append('El número de control es obligatorio')
        elif not validar_numero_control(num):
            errores.append('El número de control debe tener entre 3 y 50 caracteres alfanuméricos (puede incluir guiones y guiones bajos)')
        
        if not nombre:
            errores.append('El nombre es obligatorio')
        elif not validar_nombre(nombre):
            errores.append('El nombre debe contener solo letras y espacios, entre 2 y 150 caracteres')
        
        if not correo:
            errores.append('El correo electrónico es obligatorio para enviar el QR')
        else:
            # Depuración: log del correo recibido
            app.logger.info(f'Validando correo recibido: "{correo}"')
            resultado_validacion = validar_email(correo)
            app.logger.info(f'Resultado de validación: {resultado_validacion}')
            
            if not resultado_validacion:
                errores.append(f'El formato del correo electrónico "{correo}" no es válido. Verifique que tenga el formato correcto (ej: usuario@dominio.com)')
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            return redirect(url_for('nuevo_estudiante'))
        
        # Verificar si ya existe un estudiante con ese número de control (case-insensitive)
        existe_estudiante = Estudiante.query.filter(
            db.func.upper(Estudiante.numero_control) == num.upper()
        ).first()
        
        if existe_estudiante:
            flash(f'Ya existe un estudiante con el número de control "{existe_estudiante.numero_control}". No se pueden registrar números de control duplicados.', 'danger')
            return redirect(url_for('nuevo_estudiante'))
        
        # Verificar si ya existe un usuario con ese correo electrónico (case-insensitive)
        if correo:
            existe_usuario_correo = Usuario.query.filter(
                db.func.lower(Usuario.email) == correo.lower()
            ).first()
            
            if existe_usuario_correo:
                flash(f'Ya existe un usuario registrado con el correo electrónico "{correo}". No se pueden registrar correos duplicados.', 'danger')
                return redirect(url_for('nuevo_estudiante'))
        
        # Generar QR
        try:
            qr_base64 = generar_qr_estudiante(num)
        except Exception as e:
            flash(f'Error al generar QR: {str(e)}', 'danger')
            return redirect(url_for('nuevo_estudiante'))
        
        # Crear usuario tipo alumno automáticamente
        # Username será el número de control directamente (en mayúsculas para consistencia)
        username_alumno = num.upper()
        # Asegurar que el username sea único
        contador = 1
        username_original = username_alumno
        while Usuario.query.filter_by(username=username_alumno).first():
            username_alumno = f"{username_original}_{contador}"
            contador += 1
        
        # Crear usuario alumno
        usuario_alumno = Usuario(
            username=username_alumno,
            email=correo,
            nombre_completo=nombre,
            rol='alumno',
            password_changed=False  # Inicialmente no ha cambiado la contraseña
        )
        # Password inicial: número de control (el alumno debe cambiarlo en el primer login)
        usuario_alumno.set_password(num)
        db.session.add(usuario_alumno)
        db.session.flush()  # Para obtener el ID del usuario
        
        # Crear estudiante vinculado al usuario
        est = Estudiante(
            numero_control=num,
            nombre=nombre,
            correo=correo,
            qr_code=qr_base64,
            usuario_id=usuario_alumno.id
        )
        db.session.add(est)
        db.session.commit()
        
        # Enviar QR por correo
        if correo:
            exito, mensaje = enviar_qr_por_correo(est, qr_base64)
            if exito:
                flash(f'Estudiante registrado exitosamente. QR enviado a {correo}', 'success')
            else:
                flash(f'Estudiante registrado exitosamente, pero hubo un problema al enviar el correo: {mensaje}', 'warning')
        else:
            flash('Estudiante registrado exitosamente con QR generado', 'success')
        
        return redirect(url_for('ver_estudiantes'))
    return render_template('nuevo_estudiante.html')

# Lista estudiantes (solo admin y profesor)
@app.route('/estudiantes')
@profesor_required
def ver_estudiantes():
    usuario = get_current_user()
    if usuario.is_admin():
        estudiantes = Estudiante.query.order_by(Estudiante.nombre).all()
    else:
        # Profesor ve solo estudiantes de sus clases
        clases_profesor = Clase.query.filter_by(usuario_id=usuario.id).all()
        estudiantes_ids = set()
        for clase in clases_profesor:
            for est in clase.estudiantes:
                estudiantes_ids.add(est.id)
        estudiantes = Estudiante.query.filter(Estudiante.id.in_(estudiantes_ids)).order_by(Estudiante.nombre).all()
    return render_template('estudiantes.html', estudiantes=estudiantes, usuario=usuario)

# Eliminar estudiante
@app.route('/estudiantes/<int:estudiante_id>/eliminar', methods=['POST'])
@profesor_required
def eliminar_estudiante(estudiante_id):
    usuario = get_current_user()
    estudiante = Estudiante.query.get_or_404(estudiante_id)
    
    # Verificar permisos: admin puede eliminar cualquier estudiante, profesor solo los de sus clases
    if not usuario.is_admin():
        clases_profesor = Clase.query.filter_by(usuario_id=usuario.id).all()
        estudiantes_ids = set()
        for clase in clases_profesor:
            for est in clase.estudiantes:
                estudiantes_ids.add(est.id)
        if estudiante.id not in estudiantes_ids:
            flash('No tienes permisos para eliminar este estudiante', 'danger')
            return redirect(url_for('ver_estudiantes'))
    
    try:
        nombre_estudiante = estudiante.nombre
        numero_control = estudiante.numero_control
        
        # Eliminar usuario asociado si existe
        if estudiante.usuario_id:
            usuario_estudiante = Usuario.query.get(estudiante.usuario_id)
            if usuario_estudiante:
                db.session.delete(usuario_estudiante)
        
        # Eliminar estudiante (esto también eliminará las relaciones en estudiante_clase y asistencias por CASCADE)
        db.session.delete(estudiante)
        db.session.commit()
        
        flash(f'Estudiante {nombre_estudiante} ({numero_control}) eliminado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al eliminar estudiante: {str(e)}')
        flash('Error al eliminar el estudiante', 'danger')
    
    return redirect(url_for('ver_estudiantes'))

# Crear clase (materia) - solo admin
@app.route('/clases/nuevo', methods=['GET','POST'])
@admin_required
def nueva_clase():
    usuario = get_current_user()
    profesores = Usuario.query.filter_by(rol='profesor', activo=True).order_by(Usuario.nombre_completo).all()
    
    if request.method == 'POST':
        codigo = request.form.get('codigo_clase', '').strip()
        nombre = request.form.get('nombre_clase', '').strip()
        profesor_id = request.form.get('profesor_id', '').strip()
        
        # Validaciones
        errores = []
        
        if not codigo:
            errores.append('El código de clase es obligatorio')
        elif not validar_codigo_clase(codigo):
            errores.append('El código de clase debe tener entre 2 y 50 caracteres alfanuméricos (puede incluir guiones y guiones bajos)')
        
        if not nombre:
            errores.append('El nombre de clase es obligatorio')
        elif len(nombre) < 2 or len(nombre) > 150:
            errores.append('El nombre de clase debe tener entre 2 y 150 caracteres')
        
        if not profesor_id:
            errores.append('Debe seleccionar un profesor')
        else:
            try:
                profesor_id = int(profesor_id)
                profesor = Usuario.query.filter_by(id=profesor_id, rol='profesor', activo=True).first()
                if not profesor:
                    errores.append('El profesor seleccionado no es válido')
            except ValueError:
                errores.append('ID de profesor inválido')
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            return render_template('nueva_clase.html', profesores=profesores, codigo=codigo, nombre=nombre, profesor_id=profesor_id)
        
        # Verificar si ya existe una clase con ese código
        existe = Clase.query.filter_by(codigo_clase=codigo).first()
        if existe:
            flash('Ya existe una clase con ese código', 'warning')
            return render_template('nueva_clase.html', profesores=profesores, codigo=codigo, nombre=nombre, profesor_id=profesor_id)
        
        # Procesar horarios
        horarios_data = []
        dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
        dias_numeros = {'lunes': 1, 'martes': 2, 'miercoles': 3, 'jueves': 4, 'viernes': 5}
        
        for dia in dias_semana:
            dia_seleccionado = request.form.get(f'horario_{dia}_activo', '') == 'on'
            if dia_seleccionado:
                hora_inicio_str = request.form.get(f'horario_{dia}_inicio', '').strip()
                hora_fin_str = request.form.get(f'horario_{dia}_fin', '').strip()
                
                if not hora_inicio_str or not hora_fin_str:
                    errores.append(f'Debe especificar hora de inicio y fin para {dia.capitalize()}')
                else:
                    try:
                        hora_inicio = datetime.strptime(hora_inicio_str, '%H:%M').time()
                        hora_fin = datetime.strptime(hora_fin_str, '%H:%M').time()
                        
                        if hora_fin <= hora_inicio:
                            errores.append(f'La hora de fin debe ser mayor que la hora de inicio para {dia.capitalize()}')
                        else:
                            horarios_data.append({
                                'dia': dias_numeros[dia],
                                'hora_inicio': hora_inicio,
                                'hora_fin': hora_fin
                            })
                    except ValueError:
                        errores.append(f'Formato de hora inválido para {dia.capitalize()}. Use formato HH:MM')
        
        if not horarios_data:
            errores.append('Debe seleccionar al menos un día de la semana con su horario')
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            return render_template('nueva_clase.html', profesores=profesores, codigo=codigo, nombre=nombre, profesor_id=profesor_id)
        
        # Crear la clase
        c = Clase(codigo_clase=codigo, nombre_clase=nombre, usuario_id=profesor_id)
        db.session.add(c)
        db.session.flush()  # Para obtener el ID de la clase
        
        # Crear los horarios
        for horario_data in horarios_data:
            horario = HorarioClase(
                clase_id=c.id,
                dia_semana=horario_data['dia'],
                hora_inicio=horario_data['hora_inicio'],
                hora_fin=horario_data['hora_fin']
            )
            db.session.add(horario)
        
        db.session.commit()
        flash('Materia creada y asignada al profesor exitosamente', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('nueva_clase.html', profesores=profesores, codigo='', nombre='', profesor_id='')

# Mostrar QR almacenado del estudiante
@app.route('/qr/<int:estudiante_id>')
def mostrar_qr(estudiante_id):
    try:
        est = Estudiante.query.get_or_404(estudiante_id)
        if not est.qr_code:
            flash('QR no disponible para este estudiante', 'warning')
            return redirect(url_for('detalle_estudiante', estudiante_id=estudiante_id))
        
        # Decodificar base64 y servir como imagen
        try:
            qr_bytes = base64.b64decode(est.qr_code)
            buf = BytesIO(qr_bytes)
            buf.seek(0)
            return send_file(buf, mimetype='image/png', as_attachment=False, download_name=f'qr_{est.numero_control}.png')
        except Exception as e:
            app.logger.error(f'Error al decodificar QR para estudiante {estudiante_id}: {str(e)}')
            flash('Error al mostrar el QR. Por favor, regenere el QR.', 'danger')
            return redirect(url_for('detalle_estudiante', estudiante_id=estudiante_id))
    except Exception as e:
        app.logger.error(f'Error en mostrar_qr: {str(e)}')
        flash('Error al cargar el QR', 'danger')
        return redirect(url_for('ver_estudiantes'))

# Generar QR para estudiante y clase específica (para compatibilidad)
@app.route('/qr/<int:estudiante_id>/<int:clase_id>')
def generar_qr(estudiante_id, clase_id):
    est = Estudiante.query.get_or_404(estudiante_id)
    clase = Clase.query.get_or_404(clase_id)
    # Payload con numero_control y clase_id
    payload = {
        "numero_control": est.numero_control,
        "clase_id": clase.id
    }
    payload_str = json.dumps(payload)
    # Generar QR usando Pillow
    qr = qrcode.QRCode(box_size=10, border=2)
    qr.add_data(payload_str)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white", image_factory=PilImage)
    
    buf = BytesIO()
    img.save(buf, 'PNG')  # Formato como segundo argumento posicional
    buf.seek(0)
    # Servir como imagen
    return send_file(buf, mimetype='image/png', as_attachment=False, download_name=f'qr_{est.numero_control}_{clase.codigo_clase}.png')

# Página para ver detalle de estudiante y link para QR
@app.route('/estudiante/<int:estudiante_id>')
@login_required
def detalle_estudiante(estudiante_id):
    try:
        usuario = get_current_user()
        est = Estudiante.query.get_or_404(estudiante_id)
        
        # Si es alumno, solo puede ver su propio perfil
        if usuario.is_alumno():
            if est.usuario_id != usuario.id:
                flash('No tienes permisos para ver este perfil', 'danger')
                return redirect(url_for('alumno_dashboard'))
            clases = est.clases.all()
        else:
            # Profesor y admin pueden ver cualquier estudiante
            if usuario.is_admin():
                clases = Clase.query.order_by(Clase.nombre_clase).all()
            else:
                clases = Clase.query.filter_by(usuario_id=usuario.id).order_by(Clase.nombre_clase).all()
        
        return render_template('detalle_estudiante.html', estudiante=est, clases=clases)
    except Exception as e:
        app.logger.error(f'Error en detalle_estudiante para ID {estudiante_id}: {str(e)}')
        flash('Error al cargar el detalle del estudiante', 'danger')
        usuario = get_current_user()
        if usuario and usuario.is_alumno():
            return redirect(url_for('alumno_dashboard'))
        return redirect(url_for('ver_estudiantes'))

# Ruta para que el alumno vea su propio QR
@app.route('/alumno/mi_qr')
@alumno_required
def alumno_mi_qr():
    usuario = get_current_user()
    estudiante = Estudiante.query.filter_by(usuario_id=usuario.id).first()
    
    if not estudiante:
        flash('No se encontró información del estudiante', 'danger')
        return redirect(url_for('alumno_dashboard'))
    
    return render_template('alumno_qr.html', estudiante=estudiante)

# Cambiar contraseña (para todos los usuarios)
@app.route('/cambiar-contraseña', methods=['GET', 'POST'])
@login_required
def cambiar_contraseña():
    usuario = get_current_user()
    
    # Si el usuario ya cambió su contraseña, no permitir cambiarla de nuevo (excepto admin restableciendo)
    # Los usuarios solo pueden cambiar su contraseña una vez
    if usuario.password_changed:
        if usuario.is_alumno():
            flash('Ya has cambiado tu contraseña. Solo puedes cambiarla una vez. Si olvidaste tu contraseña, contacta al administrador.', 'warning')
            return redirect(url_for('alumno_dashboard'))
        else:
            flash('Ya has cambiado tu contraseña. Solo puedes cambiarla una vez. Si olvidaste tu contraseña, contacta al administrador.', 'warning')
            return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        password_actual = request.form.get('password_actual', '')
        password_nueva = request.form.get('password_nueva', '')
        password_confirmar = request.form.get('password_confirmar', '')
        
        errores = []
        
        if not password_actual:
            errores.append('La contraseña actual es obligatoria')
        elif not usuario.check_password(password_actual):
            errores.append('La contraseña actual es incorrecta')
        
        if not password_nueva:
            errores.append('La nueva contraseña es obligatoria')
        elif len(password_nueva) < 6:
            errores.append('La nueva contraseña debe tener al menos 6 caracteres')
        
        if password_nueva != password_confirmar:
            errores.append('Las contraseñas nuevas no coinciden')
        
        if password_actual == password_nueva:
            errores.append('La nueva contraseña debe ser diferente a la actual')
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            return redirect(url_for('cambiar_contraseña'))
        
        # Actualizar contraseña
        try:
            usuario.set_password(password_nueva)
            # Marcar que el usuario ya cambió su contraseña (aplica para todos los roles)
            usuario.password_changed = True
            db.session.commit()
            flash('Contraseña actualizada exitosamente', 'success')
            
            # Redirigir según el rol
            if usuario.is_alumno():
                return redirect(url_for('alumno_dashboard'))
            else:
                return redirect(url_for('dashboard'))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error al cambiar contraseña: {str(e)}')
            flash('Error al actualizar la contraseña', 'danger')
            return redirect(url_for('cambiar_contraseña'))
    
    return render_template('cambiar_contraseña.html', usuario=usuario)

# Ver detalle de materia y gestionar estudiantes
@app.route('/materia/<int:clase_id>')
@profesor_required
def detalle_materia(clase_id):
    usuario = get_current_user()
    if usuario.is_admin():
        clase = Clase.query.get_or_404(clase_id)
    else:
        clase = Clase.query.filter_by(id=clase_id, usuario_id=usuario.id).first_or_404()
    estudiantes_asignados = clase.estudiantes.all()
    todos_estudiantes = Estudiante.query.order_by(Estudiante.nombre).all()
    # Filtrar estudiantes no asignados
    estudiantes_no_asignados = [e for e in todos_estudiantes if e not in estudiantes_asignados]
    # Obtener horarios de la clase
    horarios = clase.horarios.order_by(HorarioClase.dia_semana).all()
    # Obtener el día de la semana actual (1=Lunes, 2=Martes, ..., 5=Viernes)
    dia_semana_actual = date.today().weekday() + 1
    # Obtener el horario de hoy si existe
    horario_hoy = None
    for horario in horarios:
        if horario.dia_semana == dia_semana_actual:
            horario_hoy = horario
            break
    
    # Verificar si estamos dentro del horario permitido para pase de lista
    # Permitir desde 15 minutos antes del inicio hasta 5 minutos después del inicio
    puede_registrar = False
    motivo_bloqueo = None
    ahora = datetime.now().time()
    
    if not horario_hoy:
        puede_registrar = False
        motivo_bloqueo = "No hay clase programada para hoy"
    else:
        # Calcular hora mínima (15 minutos antes del inicio)
        hora_inicio_dt = datetime.combine(date.today(), horario_hoy.hora_inicio)
        hora_minima_dt = hora_inicio_dt - timedelta(minutes=15)
        hora_minima = hora_minima_dt.time()
        
        # Calcular hora máxima (5 minutos después del inicio)
        hora_maxima_dt = hora_inicio_dt + timedelta(minutes=5)
        hora_maxima = hora_maxima_dt.time()
        
        if ahora < hora_minima:
            puede_registrar = False
            motivo_bloqueo = f"El pase de lista está disponible desde las {hora_minima.strftime('%H:%M')} (15 min antes del inicio)"
        elif ahora > hora_maxima:
            puede_registrar = False
            motivo_bloqueo = f"El pase de lista cerró a las {hora_maxima.strftime('%H:%M')} (5 min después del inicio)"
        else:
            puede_registrar = True
    
    # Calcular horas de pase de lista para el template
    hora_minima_pase = None
    hora_maxima_pase = None
    if horario_hoy:
        hora_inicio_dt = datetime.combine(date.today(), horario_hoy.hora_inicio)
        hora_minima_dt = hora_inicio_dt - timedelta(minutes=15)
        hora_minima_pase = hora_minima_dt.time()
        hora_maxima_dt = hora_inicio_dt + timedelta(minutes=5)
        hora_maxima_pase = hora_maxima_dt.time()
    
    return render_template('detalle_materia.html', 
                         clase=clase, 
                         estudiantes_asignados=estudiantes_asignados,
                         estudiantes_no_asignados=estudiantes_no_asignados,
                         horarios=horarios,
                         horario_hoy=horario_hoy,
                         dia_semana_actual=dia_semana_actual,
                         puede_registrar=puede_registrar,
                         motivo_bloqueo=motivo_bloqueo,
                         hora_actual=ahora,
                         hora_minima_pase=hora_minima_pase,
                         hora_maxima_pase=hora_maxima_pase)

# Asignar estudiante a materia
@app.route('/materia/<int:clase_id>/asignar', methods=['POST'])
@profesor_required
def asignar_estudiante(clase_id):
    usuario = get_current_user()
    if usuario.is_admin():
        clase = Clase.query.get_or_404(clase_id)
    else:
        clase = Clase.query.filter_by(id=clase_id, usuario_id=usuario.id).first_or_404()
    estudiante_id = request.form.get('estudiante_id')
    
    if not estudiante_id:
        flash('Debe seleccionar un estudiante', 'warning')
        return redirect(url_for('detalle_materia', clase_id=clase_id))
    
    try:
        estudiante_id = int(estudiante_id)
        estudiante = Estudiante.query.get_or_404(estudiante_id)
        
        # Verificar si ya está asignado
        if estudiante in clase.estudiantes.all():
            flash('El estudiante ya está asignado a esta materia', 'warning')
        else:
            clase.estudiantes.append(estudiante)
            db.session.commit()
            flash(f'Estudiante {estudiante.nombre} asignado exitosamente', 'success')
    except (ValueError, TypeError):
        flash('ID de estudiante inválido', 'danger')
    except Exception as e:
        app.logger.error(f'Error al asignar estudiante: {str(e)}')
        flash('Error al asignar estudiante', 'danger')
    
    return redirect(url_for('detalle_materia', clase_id=clase_id))

# Remover estudiante de materia
@app.route('/materia/<int:clase_id>/remover/<int:estudiante_id>', methods=['POST'])
@profesor_required
def remover_estudiante(clase_id, estudiante_id):
    usuario = get_current_user()
    if usuario.is_admin():
        clase = Clase.query.get_or_404(clase_id)
    else:
        clase = Clase.query.filter_by(id=clase_id, usuario_id=usuario.id).first_or_404()
    estudiante = Estudiante.query.get_or_404(estudiante_id)
    
    try:
        if estudiante in clase.estudiantes.all():
            clase.estudiantes.remove(estudiante)
            db.session.commit()
            flash(f'Estudiante {estudiante.nombre} removido de la materia', 'success')
        else:
            flash('El estudiante no está asignado a esta materia', 'warning')
    except Exception as e:
        app.logger.error(f'Error al remover estudiante: {str(e)}')
        flash('Error al remover estudiante', 'danger')
    
    return redirect(url_for('detalle_materia', clase_id=clase_id))

# Eliminar materia
@app.route('/materia/<int:clase_id>/eliminar', methods=['POST'])
@profesor_required
def eliminar_materia(clase_id):
    usuario = get_current_user()
    if usuario.is_admin():
        clase = Clase.query.get_or_404(clase_id)
    else:
        clase = Clase.query.filter_by(id=clase_id, usuario_id=usuario.id).first_or_404()
    
    try:
        nombre_clase = clase.nombre_clase
        db.session.delete(clase)
        db.session.commit()
        flash(f'Materia "{nombre_clase}" eliminada exitosamente', 'success')
    except Exception as e:
        app.logger.error(f'Error al eliminar materia: {str(e)}')
        flash('Error al eliminar materia', 'danger')
    
    return redirect(url_for('dashboard'))

# Escáner - vista web para leer QR con cámara del móvil
@app.route('/scanner/<int:clase_id>')
@profesor_required
def scanner(clase_id):
    usuario = get_current_user()
    if usuario.is_admin():
        clase = Clase.query.get_or_404(clase_id)
    else:
        clase = Clase.query.filter_by(id=clase_id, usuario_id=usuario.id).first_or_404()
    return render_template('scanner.html', clase=clase)

# Endpoint que recibe el contenido del QR y registra la asistencia
@app.route('/registrar_asistencia', methods=['POST'])
def registrar_asistencia():
    data = request.get_json() or {}
    
    # El QR puede contener solo numero_control o numero_control + clase_id
    numero = data.get('numero_control')
    clase_id = data.get('clase_id')
    
    if not numero:
        return jsonify({"ok": False, "msg": "Número de control no encontrado en el QR"}), 400
    
    # Validar formato del número de control
    if not validar_numero_control(numero):
        return jsonify({"ok": False, "msg": "Formato de número de control inválido"}), 400
    
    estudiante = Estudiante.query.filter_by(numero_control=numero).first()
    if not estudiante:
        return jsonify({"ok": False, "msg": "Estudiante no registrado"}), 404
    
    # Si no viene clase_id en el QR, usar la clase del scanner
    if not clase_id:
        return jsonify({"ok": False, "msg": "Clase no especificada"}), 400
    
    try:
        clase_id = int(clase_id)
    except (ValueError, TypeError):
        return jsonify({"ok": False, "msg": "ID de clase inválido"}), 400
    
    clase = Clase.query.get(clase_id)
    if not clase:
        return jsonify({"ok": False, "msg": "Clase no encontrada"}), 404

    # Verificar que el estudiante esté asignado a la clase
    # Usar una consulta más eficiente para verificar la relación
    estudiante_en_clase = db.session.query(estudiante_clase).filter_by(
        estudiante_id=estudiante.id,
        clase_id=clase.id
    ).first()
    
    if not estudiante_en_clase:
        return jsonify({
            "ok": False, 
            "msg": f"El estudiante {estudiante.nombre} ({estudiante.numero_control}) no está asignado a la materia '{clase.nombre_clase}'. Solo se pueden registrar asistencias de estudiantes asignados a esta clase.",
            "estudiante": estudiante.nombre,
            "numero_control": estudiante.numero_control,
            "clase": clase.nombre_clase
        }), 403

    hoy = date.today()
    ahora = datetime.now().time()
    
    # Obtener el día de la semana (0=Lunes, 1=Martes, ..., 4=Viernes)
    dia_semana_actual = hoy.weekday() + 1  # weekday() retorna 0-6, necesitamos 1-5
    
    # Verificar si hay horario configurado para este día
    horario_hoy = HorarioClase.query.filter_by(
        clase_id=clase.id,
        dia_semana=dia_semana_actual
    ).first()
    
    if not horario_hoy:
        dias_nombres = {1: 'Lunes', 2: 'Martes', 3: 'Miércoles', 4: 'Jueves', 5: 'Viernes'}
        return jsonify({
            "ok": False,
            "msg": f"No hay clase programada para {dias_nombres.get(dia_semana_actual, 'este día')}. La materia '{clase.nombre_clase}' no tiene horario configurado para este día.",
            "estudiante": estudiante.nombre,
            "clase": clase.nombre_clase
        }), 400
    
    # Validar que la hora actual esté dentro del intervalo permitido para pase de lista
    # Permitir desde 15 minutos antes del inicio hasta 5 minutos después del inicio
    hora_inicio_dt = datetime.combine(hoy, horario_hoy.hora_inicio)
    hora_minima_dt = hora_inicio_dt - timedelta(minutes=15)
    hora_minima = hora_minima_dt.time()
    
    hora_maxima_dt = hora_inicio_dt + timedelta(minutes=5)
    hora_maxima = hora_maxima_dt.time()
    
    if ahora < hora_minima:
        return jsonify({
            "ok": False,
            "msg": f"Fuera del horario de pase de lista. El pase de lista está disponible desde las {hora_minima.strftime('%H:%M')} (15 minutos antes del inicio de clase). Hora actual: {ahora.strftime('%H:%M')}",
            "estudiante": estudiante.nombre,
            "clase": clase.nombre_clase,
            "horario": f"Pase de lista: {hora_minima.strftime('%H:%M')} - {hora_maxima.strftime('%H:%M')}"
        }), 400
    
    if ahora > hora_maxima:
        return jsonify({
            "ok": False,
            "msg": f"Fuera del horario de pase de lista. El pase de lista cerró a las {hora_maxima.strftime('%H:%M')} (5 minutos después del inicio de clase). Hora actual: {ahora.strftime('%H:%M')}",
            "estudiante": estudiante.nombre,
            "clase": clase.nombre_clase,
            "horario": f"Pase de lista: {hora_minima.strftime('%H:%M')} - {hora_maxima.strftime('%H:%M')}"
        }), 400
    
    # Verificar si ya tiene asistencia hoy
    existe = Asistencia.query.filter_by(estudiante_id=estudiante.id, clase_id=clase.id, fecha=hoy).first()
    if existe:
        return jsonify({
            "ok": False, 
            "msg": f"Asistencia ya registrada hoy a las {existe.hora.strftime('%H:%M:%S')}",
            "estudiante": estudiante.nombre
        }), 409

    nueva = Asistencia(estudiante_id=estudiante.id, clase_id=clase.id, fecha=hoy, hora=ahora, metodo='QR')
    db.session.add(nueva)
    db.session.commit()
    return jsonify({
        "ok": True, 
        "msg": "Asistencia registrada exitosamente", 
        "estudiante": estudiante.nombre,
        "hora": ahora.strftime('%H:%M:%S')
    })

# Regenerar QR para un estudiante existente
@app.route('/estudiante/<int:estudiante_id>/regenerar_qr', methods=['POST'])
@login_required
def regenerar_qr(estudiante_id):
    est = Estudiante.query.get_or_404(estudiante_id)
    try:
        qr_base64 = generar_qr_estudiante(est.numero_control)
        est.qr_code = qr_base64
        db.session.commit()
        flash('QR regenerado exitosamente', 'success')
    except Exception as e:
        flash(f'Error al regenerar QR: {str(e)}', 'danger')
    return redirect(url_for('detalle_estudiante', estudiante_id=estudiante_id))

# Ver asistencias (simple)
@app.route('/asistencias')
@login_required
def ver_asistencias():
    usuario = get_current_user()
    clase_id = request.args.get('clase_id', type=int)
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    
    # Query base
    query = db.session.query(Asistencia, Estudiante, Clase)\
        .join(Estudiante, Estudiante.id==Asistencia.estudiante_id)\
        .join(Clase, Clase.id==Asistencia.clase_id)\
        .filter(Clase.usuario_id == usuario.id)
    
    # Filtros opcionales
    if clase_id:
        query = query.filter(Clase.id == clase_id)
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            query = query.filter(Asistencia.fecha >= fecha_desde_obj)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            query = query.filter(Asistencia.fecha <= fecha_hasta_obj)
        except ValueError:
            pass
    
    rows = query.order_by(Asistencia.fecha.desc(), Asistencia.hora.desc()).all()
    
    # Obtener todas las clases del usuario para el filtro
    clases = Clase.query.filter_by(usuario_id=usuario.id).order_by(Clase.nombre_clase).all()
    
    return render_template('asistencias.html', rows=rows, clases=clases, usuario=usuario,
                         clase_id=clase_id, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)

# Función para generar reporte Excel de una clase específica del día actual
def generar_reporte_excel_clase_dia(clase_id, fecha_clase):
    """Genera un reporte Excel de asistencias para una clase específica del día indicado"""
    try:
        clase = Clase.query.get(clase_id)
        if not clase:
            return None, "Clase no encontrada"
        
        profesor = Usuario.query.get(clase.usuario_id)
        if not profesor:
            return None, "Profesor no encontrado"
        
        # Obtener asistencias del día específico para esta clase
        query = db.session.query(Asistencia, Estudiante)\
            .join(Estudiante, Estudiante.id == Asistencia.estudiante_id)\
            .filter(Asistencia.clase_id == clase_id)\
            .filter(Asistencia.fecha == fecha_clase)
        
        rows = query.order_by(Asistencia.hora.asc()).all()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Asistencias"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Encabezado del reporte
        ws.merge_cells('A1:H1')
        ws['A1'] = f'REPORTE DE ASISTENCIAS - {clase.nombre_clase.upper()}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f'Fecha: {fecha_clase.strftime("%d/%m/%Y")} | Profesor: {profesor.nombre_completo}'
        ws['A2'].alignment = center_alignment
        ws['A2'].font = Font(italic=True)
        
        ws.merge_cells('A3:H3')
        fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        ws['A3'] = f'Generado el: {fecha_generacion}'
        ws['A3'].alignment = center_alignment
        ws['A3'].font = Font(italic=True, size=10)
        
        row_start = 5
        
        # Encabezados de columnas
        headers = [
            'Número de Control',
            'Nombre del Estudiante',
            'Correo',
            'Hora de Asistencia',
            'Método',
            'Estado'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_start, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment
        
        # Obtener todos los estudiantes asignados a la clase
        estudiantes_asignados = clase.estudiantes.all()
        estudiantes_dict = {est.id: est for est in estudiantes_asignados}
        asistencias_dict = {row[0].estudiante_id: row[0] for row in rows}
        
        # Datos: mostrar todos los estudiantes asignados, marcando quién asistió
        row_idx = row_start + 1
        for estudiante in estudiantes_asignados:
            asistencia = asistencias_dict.get(estudiante.id)
            
            ws.cell(row=row_idx, column=1, value=estudiante.numero_control).border = border
            ws.cell(row=row_idx, column=2, value=estudiante.nombre).border = border
            ws.cell(row=row_idx, column=3, value=estudiante.correo or 'N/A').border = border
            
            if asistencia:
                ws.cell(row=row_idx, column=4, value=asistencia.hora.strftime('%H:%M:%S')).border = border
                ws.cell(row=row_idx, column=5, value=asistencia.metodo or 'QR').border = border
                ws.cell(row=row_idx, column=6, value='Presente').border = border
                ws.cell(row=row_idx, column=6).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws.cell(row=row_idx, column=4, value='N/A').border = border
                ws.cell(row=row_idx, column=5, value='N/A').border = border
                ws.cell(row=row_idx, column=6, value='Ausente').border = border
                ws.cell(row=row_idx, column=6).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row_idx += 1
        
        # Ajustar ancho de columnas
        column_widths = [20, 35, 30, 18, 12, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Resumen al final
        total_estudiantes = len(estudiantes_asignados)
        total_asistencias = len(rows)
        total_ausentes = total_estudiantes - total_asistencias
        
        row_summary = row_idx + 2
        ws.merge_cells(f'A{row_summary}:H{row_summary}')
        ws[f'A{row_summary}'] = f'RESUMEN: Total Estudiantes: {total_estudiantes} | Presentes: {total_asistencias} | Ausentes: {total_ausentes}'
        ws[f'A{row_summary}'].font = Font(bold=True, size=12)
        ws[f'A{row_summary}'].alignment = center_alignment
        ws[f'A{row_summary}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        nombre_archivo = f'Reporte_Asistencias_{clase.codigo_clase}_{fecha_clase.strftime("%Y%m%d")}.xlsx'
        
        return output, nombre_archivo
    except Exception as e:
        app.logger.error(f'Error al generar reporte Excel: {str(e)}')
        return None, f"Error al generar reporte: {str(e)}"

# Función para enviar reporte por correo al profesor
def enviar_reporte_por_correo(clase_id, fecha_clase):
    """Envía el reporte de asistencia del día por correo al profesor"""
    try:
        clase = Clase.query.get(clase_id)
        if not clase:
            app.logger.error(f'Clase {clase_id} no encontrada para enviar reporte')
            return False
        
        profesor = Usuario.query.get(clase.usuario_id)
        if not profesor or not profesor.email:
            app.logger.error(f'Profesor de clase {clase_id} no tiene correo configurado')
            return False
        
        # Verificar si ya se envió el reporte para este día
        reporte_existente = ReporteEnviado.query.filter_by(
            clase_id=clase_id,
            fecha_clase=fecha_clase
        ).first()
        
        if reporte_existente:
            app.logger.info(f'Reporte ya enviado para clase {clase_id} el {fecha_clase}')
            return True  # Ya se envió, no es un error
        
        # Generar el reporte Excel
        output, nombre_archivo = generar_reporte_excel_clase_dia(clase_id, fecha_clase)
        if not output:
            app.logger.error(f'No se pudo generar el reporte para clase {clase_id}')
            return False
        
        # Verificar configuración de correo
        mail_password = app.config.get('MAIL_PASSWORD', '') or os.getenv('MAIL_PASSWORD', '')
        mail_username = app.config.get('MAIL_USERNAME', '') or os.getenv('MAIL_USERNAME', '')
        
        if not mail_password or mail_password.strip() == '':
            app.logger.warning('MAIL_PASSWORD no configurada, no se puede enviar reporte')
            return False
        
        # Crear el mensaje de correo
        año_actual = datetime.now().year
        msg = Message(
            subject=f'Reporte de Asistencias - {clase.nombre_clase} - {fecha_clase.strftime("%d/%m/%Y")}',
            recipients=[profesor.email],
            html=f'''
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2>Reporte de Asistencias</h2>
                <p>Estimado/a <strong>{profesor.nombre_completo}</strong>,</p>
                <p>Se adjunta el reporte de asistencias de la materia <strong>{clase.nombre_clase}</strong> ({clase.codigo_clase}) correspondiente al día <strong>{fecha_clase.strftime("%d/%m/%Y")}</strong>.</p>
                <p>El tiempo de pase de lista ha finalizado y este es el reporte automático de las asistencias registradas.</p>
                <hr>
                <p><small>Este es un correo automático generado por EduCheck.</small></p>
                <p><small>&copy; {año_actual} EduCheck</small></p>
            </body>
            </html>
            ''',
            sender=app.config.get('MAIL_DEFAULT_SENDER', mail_username)
        )
        
        # Adjuntar el archivo Excel
        output.seek(0)
        msg.attach(
            filename=nombre_archivo,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            data=output.read()
        )
        
        # Enviar el correo
        mail.init_app(app)
        mail.send(msg)
        
        # Registrar que se envió el reporte
        reporte_enviado = ReporteEnviado(
            clase_id=clase_id,
            fecha_clase=fecha_clase
        )
        db.session.add(reporte_enviado)
        db.session.commit()
        
        app.logger.info(f'Reporte enviado exitosamente a {profesor.email} para clase {clase_id} del {fecha_clase}')
        return True
        
    except Exception as e:
        app.logger.error(f'Error al enviar reporte por correo: {str(e)}')
        db.session.rollback()
        return False

# Función que verifica y envía reportes automáticamente
def verificar_y_enviar_reportes():
    """Verifica qué clases han terminado su tiempo de pase de lista y envía reportes"""
    with app.app_context():
        try:
            hoy = date.today()
            ahora = datetime.now().time()
            dia_semana_actual = hoy.weekday() + 1  # 1=Lunes, 2=Martes, etc.
            
            # Obtener todas las clases con horario para hoy
            horarios_hoy = HorarioClase.query.filter_by(dia_semana=dia_semana_actual).all()
            
            for horario in horarios_hoy:
                clase = horario.clase
                
                # Calcular hora máxima de pase de lista (5 minutos después del inicio)
                hora_inicio_dt = datetime.combine(hoy, horario.hora_inicio)
                hora_maxima_dt = hora_inicio_dt + timedelta(minutes=5)
                hora_maxima = hora_maxima_dt.time()
                
                # Verificar si ya pasó el tiempo de pase de lista (con margen de 1 minuto)
                # Esto permite que se ejecute hasta 1 minuto después de cerrar el pase de lista
                hora_verificacion = (datetime.combine(hoy, ahora) + timedelta(minutes=1)).time()
                
                if hora_verificacion >= hora_maxima:
                    # Verificar si ya se envió el reporte para este día
                    reporte_existente = ReporteEnviado.query.filter_by(
                        clase_id=clase.id,
                        fecha_clase=hoy
                    ).first()
                    
                    if not reporte_existente:
                        # Verificar si hay asistencias registradas para esta clase hoy
                        asistencias_hoy = Asistencia.query.filter_by(
                            clase_id=clase.id,
                            fecha=hoy
                        ).count()
                        
                        # Solo enviar reporte si hay al menos una asistencia registrada
                        if asistencias_hoy > 0:
                            app.logger.info(f'Enviando reporte automático para clase {clase.id} ({clase.nombre_clase}) del {hoy} - {asistencias_hoy} asistencias registradas')
                            enviar_reporte_por_correo(clase.id, hoy)
                        else:
                            app.logger.debug(f'No se envía reporte para clase {clase.id} ({clase.nombre_clase}) del {hoy} - No hay asistencias registradas')
                    else:
                        app.logger.debug(f'Reporte ya enviado para clase {clase.id} del {hoy}')
                        
        except Exception as e:
            app.logger.error(f'Error en verificar_y_enviar_reportes: {str(e)}')

# Configurar el scheduler (se inicializará al final del archivo)
scheduler = None

# Generar reporte Excel de asistencias
@app.route('/asistencias/reporte_excel')
@login_required
def generar_reporte_excel():
    usuario = get_current_user()
    clase_id = request.args.get('clase_id', type=int)
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    
    # Query base
    query = db.session.query(Asistencia, Estudiante, Clase, Usuario)\
        .join(Estudiante, Estudiante.id==Asistencia.estudiante_id)\
        .join(Clase, Clase.id==Asistencia.clase_id)\
        .join(Usuario, Usuario.id==Clase.usuario_id)\
        .filter(Clase.usuario_id == usuario.id)
    
    # Aplicar filtros
    if clase_id:
        query = query.filter(Clase.id == clase_id)
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            query = query.filter(Asistencia.fecha >= fecha_desde_obj)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            query = query.filter(Asistencia.fecha <= fecha_hasta_obj)
        except ValueError:
            pass
    
    rows = query.order_by(Asistencia.fecha.desc(), Asistencia.hora.desc()).all()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Asistencias"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Encabezado del reporte
    ws.merge_cells('A1:K1')
    ws['A1'] = f'REPORTE DE ASISTENCIAS - {usuario.nombre_completo.upper()}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment
    
    ws.merge_cells('A2:K2')
    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws['A2'] = f'Generado el: {fecha_generacion}'
    ws['A2'].alignment = center_alignment
    ws['A2'].font = Font(italic=True)
    
    # Información de filtros aplicados
    filtros = []
    if clase_id:
        clase = Clase.query.get(clase_id)
        if clase:
            filtros.append(f"Materia: {clase.nombre_clase}")
    if fecha_desde:
        filtros.append(f"Desde: {fecha_desde}")
    if fecha_hasta:
        filtros.append(f"Hasta: {fecha_hasta}")
    
    if filtros:
        ws.merge_cells('A3:K3')
        ws['A3'] = 'Filtros aplicados: ' + ' | '.join(filtros)
        ws['A3'].font = Font(italic=True, size=10)
        ws['A3'].alignment = Alignment(horizontal='left')
        row_start = 5
    else:
        row_start = 4
    
    # Encabezados de columnas
    headers = [
        'ID Asistencia',
        'Fecha',
        'Hora',
        'Profesor',
        'Materia',
        'Código Materia',
        'Estudiante',
        'Número de Control',
        'Correo Estudiante',
        'Método',
        'Registrado el'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_start, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment
    
    # Datos
    for row_idx, (asistencia, estudiante, clase, profesor) in enumerate(rows, start=row_start + 1):
        ws.cell(row=row_idx, column=1, value=asistencia.id).border = border
        ws.cell(row=row_idx, column=2, value=asistencia.fecha.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row_idx, column=3, value=asistencia.hora.strftime('%H:%M:%S')).border = border
        ws.cell(row=row_idx, column=4, value=profesor.nombre_completo).border = border
        ws.cell(row=row_idx, column=5, value=clase.nombre_clase).border = border
        ws.cell(row=row_idx, column=6, value=clase.codigo_clase).border = border
        ws.cell(row=row_idx, column=7, value=estudiante.nombre).border = border
        ws.cell(row=row_idx, column=8, value=estudiante.numero_control).border = border
        ws.cell(row=row_idx, column=9, value=estudiante.correo or 'N/A').border = border
        ws.cell(row=row_idx, column=10, value=asistencia.metodo).border = border
        ws.cell(row=row_idx, column=11, value=asistencia.creado_en.strftime('%d/%m/%Y %H:%M:%S') if asistencia.creado_en else 'N/A').border = border
    
    # Ajustar ancho de columnas
    column_widths = [12, 12, 10, 25, 30, 15, 30, 15, 30, 10, 20]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Resumen al final
    row_summary = row_start + len(rows) + 2
    ws.merge_cells(f'A{row_summary}:K{row_summary}')
    ws[f'A{row_summary}'] = f'TOTAL DE ASISTENCIAS: {len(rows)}'
    ws[f'A{row_summary}'].font = Font(bold=True, size=12)
    ws[f'A{row_summary}'].alignment = center_alignment
    ws[f'A{row_summary}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre del archivo
    fecha_archivo = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f'Reporte_Asistencias_{fecha_archivo}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo
    )

# ---------------------
# RUTAS DE ADMINISTRACIÓN (solo admin)
# ---------------------

# Gestión de administradores
@app.route('/admin/administradores')
@admin_required
def admin_administradores():
    administradores = Usuario.query.filter_by(rol='admin').order_by(Usuario.nombre_completo).all()
    return render_template('admin_administradores.html', administradores=administradores)

# Eliminar administrador (admin) - Requiere contraseña de confirmación
@app.route('/admin/administradores/<int:admin_id>/eliminar', methods=['POST'])
@admin_required
def eliminar_administrador(admin_id):
    usuario_actual = get_current_user()
    admin_a_eliminar = Usuario.query.get_or_404(admin_id)
    
    # Contraseña requerida para eliminar administradores
    PASSWORD_ELIMINAR_ADMIN = "Eliminaradm"
    
    # Obtener la contraseña del formulario
    password_confirmacion = request.form.get('password_confirmacion', '').strip()
    
    # Verificar contraseña de confirmación
    if password_confirmacion != PASSWORD_ELIMINAR_ADMIN:
        flash('Contraseña de confirmación incorrecta. No se pudo eliminar el administrador.', 'danger')
        return redirect(url_for('admin_administradores'))
    
    # Verificar que el usuario a eliminar sea realmente un administrador
    if admin_a_eliminar.rol != 'admin':
        flash('El usuario seleccionado no es un administrador', 'danger')
        return redirect(url_for('admin_administradores'))
    
    # No permitir que un admin se elimine a sí mismo
    if admin_a_eliminar.id == usuario_actual.id:
        flash('No puedes eliminar tu propia cuenta de administrador', 'danger')
        return redirect(url_for('admin_administradores'))
    
    # Verificar que no sea el último administrador
    total_admins = Usuario.query.filter_by(rol='admin').count()
    if total_admins <= 1:
        flash('No se puede eliminar el último administrador. Debe haber al menos un administrador en el sistema.', 'danger')
        return redirect(url_for('admin_administradores'))
    
    try:
        nombre_admin = admin_a_eliminar.nombre_completo
        username_admin = admin_a_eliminar.username
        
        # Si el admin tiene clases asignadas, se eliminarán por CASCADE
        # Eliminar el usuario (esto también eliminará las clases por CASCADE)
        db.session.delete(admin_a_eliminar)
        db.session.commit()
        
        flash(f'Administrador {nombre_admin} ({username_admin}) eliminado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al eliminar administrador: {str(e)}')
        flash('Error al eliminar el administrador', 'danger')
    
    return redirect(url_for('admin_administradores'))

# Restablecer contraseña de usuario (solo admin)
@app.route('/admin/usuarios/<int:usuario_id>/restablecer-contraseña', methods=['GET', 'POST'])
@admin_required
def admin_restablecer_contraseña(usuario_id):
    usuario_objetivo = Usuario.query.get_or_404(usuario_id)
    usuario_actual = get_current_user()
    
    if request.method == 'POST':
        nueva_password = request.form.get('nueva_password', '').strip()
        confirmar_password = request.form.get('confirmar_password', '').strip()
        
        errores = []
        
        if not nueva_password:
            errores.append('La nueva contraseña es obligatoria')
        elif len(nueva_password) < 6:
            errores.append('La contraseña debe tener al menos 6 caracteres')
        
        if nueva_password != confirmar_password:
            errores.append('Las contraseñas no coinciden')
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            return render_template('admin_restablecer_contraseña.html', 
                                 usuario_objetivo=usuario_objetivo,
                                 usuario_actual=usuario_actual)
        
        try:
            # Restablecer la contraseña
            usuario_objetivo.set_password(nueva_password)
            # Al restablecer la contraseña, el usuario debe cambiarla en su próximo login (aplica para todos los roles)
            usuario_objetivo.password_changed = False
            db.session.commit()
            
            flash(f'Contraseña restablecida exitosamente para {usuario_objetivo.nombre_completo} ({usuario_objetivo.username}). El usuario deberá cambiarla en su próximo inicio de sesión.', 'success')
            
            # Redirigir según el tipo de usuario
            if usuario_objetivo.rol == 'admin':
                return redirect(url_for('admin_administradores'))
            elif usuario_objetivo.rol == 'profesor':
                return redirect(url_for('admin_profesores'))
            elif usuario_objetivo.rol == 'alumno':
                return redirect(url_for('ver_estudiantes'))
            else:
                return redirect(url_for('dashboard'))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error al restablecer contraseña: {str(e)}')
            flash('Error al restablecer la contraseña', 'danger')
            return render_template('admin_restablecer_contraseña.html', 
                                 usuario_objetivo=usuario_objetivo,
                                 usuario_actual=usuario_actual)
    
    return render_template('admin_restablecer_contraseña.html', 
                         usuario_objetivo=usuario_objetivo,
                         usuario_actual=usuario_actual)

# Crear administrador (admin)
@app.route('/admin/administradores/nuevo', methods=['GET', 'POST'])
@admin_required
def admin_nuevo_administrador():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        password_confirm = request.form.get('password_confirm', '')
        nombre_completo = request.form.get('nombre_completo', '').strip()
        
        errores = []
        
        if not username:
            errores.append('El nombre de usuario es obligatorio')
        elif not validar_username(username):
            errores.append('El nombre de usuario debe tener entre 3 y 30 caracteres alfanuméricos')
        
        if not email:
            errores.append('El correo electrónico es obligatorio')
        elif not validar_email(email):
            errores.append('El formato del correo electrónico no es válido')
        
        if not password:
            errores.append('La contraseña es obligatoria')
        elif len(password) < 6:
            errores.append('La contraseña debe tener al menos 6 caracteres')
        
        if password != password_confirm:
            errores.append('Las contraseñas no coinciden')
        
        if not nombre_completo:
            errores.append('El nombre completo es obligatorio')
        elif not validar_nombre(nombre_completo):
            errores.append('El nombre debe contener solo letras y espacios')
        
        if Usuario.query.filter_by(username=username).first():
            errores.append('El nombre de usuario ya está en uso')
        
        if Usuario.query.filter_by(email=email).first():
            errores.append('El correo electrónico ya está registrado')
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            return redirect(url_for('admin_nuevo_administrador'))
        
        nuevo_admin = Usuario(
            username=username,
            email=email,
            nombre_completo=nombre_completo,
            rol='admin',
            password_changed=False  # Debe cambiar la contraseña en el primer login
        )
        nuevo_admin.set_password(password)
        
        db.session.add(nuevo_admin)
        db.session.commit()
        
        # Enviar correo de bienvenida
        try:
            enviado, mensaje = enviar_correo_bienvenida(nuevo_admin, password)
            if enviado:
                flash(f'Administrador {nombre_completo} creado exitosamente. Se envió correo de bienvenida.', 'success')
            else:
                flash(f'Administrador {nombre_completo} creado exitosamente, pero hubo un problema al enviar el correo: {mensaje}', 'warning')
        except Exception as e:
            app.logger.error(f'Error al enviar correo de bienvenida: {str(e)}')
            flash(f'Administrador {nombre_completo} creado exitosamente, pero hubo un problema al enviar el correo.', 'warning')
        
        return redirect(url_for('admin_administradores'))
    
    return render_template('admin_nuevo_administrador.html')

# Gestión de profesores
@app.route('/admin/profesores')
@admin_required
def admin_profesores():
    profesores = Usuario.query.filter_by(rol='profesor').order_by(Usuario.nombre_completo).all()
    return render_template('admin_profesores.html', profesores=profesores)

# Crear profesor (admin)
@app.route('/admin/profesores/nuevo', methods=['GET', 'POST'])
@admin_required
def admin_nuevo_profesor():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        password_confirm = request.form.get('password_confirm', '')
        nombre_completo = request.form.get('nombre_completo', '').strip()
        
        errores = []
        
        if not username:
            errores.append('El nombre de usuario es obligatorio')
        elif not validar_username(username):
            errores.append('El nombre de usuario debe tener entre 3 y 30 caracteres alfanuméricos')
        
        if not email:
            errores.append('El correo electrónico es obligatorio')
        elif not validar_email(email):
            errores.append('El formato del correo electrónico no es válido')
        
        if not password:
            errores.append('La contraseña es obligatoria')
        elif len(password) < 6:
            errores.append('La contraseña debe tener al menos 6 caracteres')
        
        if password != password_confirm:
            errores.append('Las contraseñas no coinciden')
        
        if not nombre_completo:
            errores.append('El nombre completo es obligatorio')
        elif not validar_nombre(nombre_completo):
            errores.append('El nombre debe contener solo letras y espacios')
        
        if Usuario.query.filter_by(username=username).first():
            errores.append('El nombre de usuario ya está en uso')
        
        if Usuario.query.filter_by(email=email).first():
            errores.append('El correo electrónico ya está registrado')
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            return redirect(url_for('admin_nuevo_profesor'))
        
        nuevo_profesor = Usuario(
            username=username,
            email=email,
            nombre_completo=nombre_completo,
            rol='profesor',
            password_changed=False  # Debe cambiar la contraseña en el primer login
        )
        nuevo_profesor.set_password(password)
        
        db.session.add(nuevo_profesor)
        db.session.commit()
        
        # Enviar correo de bienvenida
        try:
            enviado, mensaje = enviar_correo_bienvenida(nuevo_profesor, password)
            if enviado:
                flash(f'Profesor {nombre_completo} creado exitosamente. Se envió correo de bienvenida.', 'success')
            else:
                flash(f'Profesor {nombre_completo} creado exitosamente, pero hubo un problema al enviar el correo: {mensaje}', 'warning')
        except Exception as e:
            app.logger.error(f'Error al enviar correo de bienvenida: {str(e)}')
            flash(f'Profesor {nombre_completo} creado exitosamente, pero hubo un problema al enviar el correo.', 'warning')
        
        return redirect(url_for('admin_profesores'))
    
    return render_template('admin_nuevo_profesor.html')

# Asignar estudiantes a profesores (admin)
@app.route('/admin/asignar')
@admin_required
def admin_asignar():
    profesores = Usuario.query.filter_by(rol='profesor', activo=True).order_by(Usuario.nombre_completo).all()
    estudiantes = Estudiante.query.order_by(Estudiante.nombre).all()
    clases = Clase.query.order_by(Clase.nombre_clase).all()
    
    return render_template('admin_asignar.html',
                         profesores=profesores,
                         estudiantes=estudiantes,
                         clases=clases)

# Reasignar materia a otro profesor (admin)
@app.route('/admin/materia/<int:clase_id>/reasignar', methods=['GET', 'POST'])
@admin_required
def reasignar_materia(clase_id):
    clase = Clase.query.get_or_404(clase_id)
    profesores = Usuario.query.filter_by(rol='profesor', activo=True).order_by(Usuario.nombre_completo).all()
    
    if request.method == 'POST':
        nuevo_profesor_id = request.form.get('profesor_id', '').strip()
        
        if not nuevo_profesor_id:
            flash('Debe seleccionar un profesor', 'danger')
            return redirect(url_for('reasignar_materia', clase_id=clase_id))
        
        try:
            nuevo_profesor_id = int(nuevo_profesor_id)
            nuevo_profesor = Usuario.query.filter_by(id=nuevo_profesor_id, rol='profesor', activo=True).first()
            
            if not nuevo_profesor:
                flash('El profesor seleccionado no es válido', 'danger')
                return redirect(url_for('reasignar_materia', clase_id=clase_id))
            
            profesor_anterior = clase.maestro.nombre_completo if clase.maestro else 'Sin asignar'
            clase.usuario_id = nuevo_profesor_id
            db.session.commit()
            
            flash(f'Materia "{clase.nombre_clase}" reasignada de {profesor_anterior} a {nuevo_profesor.nombre_completo}', 'success')
            return redirect(url_for('admin_asignar'))
        except ValueError:
            flash('ID de profesor inválido', 'danger')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error al reasignar materia: {str(e)}')
            flash('Error al reasignar la materia', 'danger')
    
    return render_template('reasignar_materia.html', clase=clase, profesores=profesores)

# Editar horarios de una materia (admin)
@app.route('/admin/materia/<int:clase_id>/horarios', methods=['GET', 'POST'])
@admin_required
def editar_horarios(clase_id):
    clase = Clase.query.get_or_404(clase_id)
    
    if request.method == 'POST':
        # Obtener horarios actuales
        horarios_actuales = {h.dia_semana: h for h in clase.horarios.all()}
        
        # Procesar horarios del formulario
        horarios_data = []
        dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
        dias_numeros = {'lunes': 1, 'martes': 2, 'miercoles': 3, 'jueves': 4, 'viernes': 5}
        
        errores = []
        
        for dia in dias_semana:
            dia_seleccionado = request.form.get(f'horario_{dia}_activo', '') == 'on'
            dia_numero = dias_numeros[dia]
            
            if dia_seleccionado:
                hora_inicio_str = request.form.get(f'horario_{dia}_inicio', '').strip()
                hora_fin_str = request.form.get(f'horario_{dia}_fin', '').strip()
                
                if not hora_inicio_str or not hora_fin_str:
                    errores.append(f'Debe especificar hora de inicio y fin para {dia.capitalize()}')
                else:
                    try:
                        hora_inicio = datetime.strptime(hora_inicio_str, '%H:%M').time()
                        hora_fin = datetime.strptime(hora_fin_str, '%H:%M').time()
                        
                        if hora_fin <= hora_inicio:
                            errores.append(f'La hora de fin debe ser mayor que la hora de inicio para {dia.capitalize()}')
                        else:
                            horarios_data.append({
                                'dia': dia_numero,
                                'hora_inicio': hora_inicio,
                                'hora_fin': hora_fin
                            })
                    except ValueError:
                        errores.append(f'Formato de hora inválido para {dia.capitalize()}. Use formato HH:MM')
            else:
                # Si no está seleccionado, eliminar el horario si existe
                if dia_numero in horarios_actuales:
                    db.session.delete(horarios_actuales[dia_numero])
        
        if not horarios_data:
            errores.append('Debe seleccionar al menos un día de la semana con su horario')
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            # Recargar horarios para mostrar en el formulario
            horarios_actuales = {h.dia_semana: h for h in clase.horarios.all()}
            return render_template('editar_horarios.html', clase=clase, horarios_actuales=horarios_actuales)
        
        # Actualizar o crear horarios
        for horario_data in horarios_data:
            if horario_data['dia'] in horarios_actuales:
                # Actualizar horario existente
                horario = horarios_actuales[horario_data['dia']]
                horario.hora_inicio = horario_data['hora_inicio']
                horario.hora_fin = horario_data['hora_fin']
            else:
                # Crear nuevo horario
                horario = HorarioClase(
                    clase_id=clase.id,
                    dia_semana=horario_data['dia'],
                    hora_inicio=horario_data['hora_inicio'],
                    hora_fin=horario_data['hora_fin']
                )
                db.session.add(horario)
        
        db.session.commit()
        flash('Horarios actualizados exitosamente', 'success')
        return redirect(url_for('detalle_materia', clase_id=clase_id))
    
    # GET: Mostrar formulario con horarios actuales
    horarios_actuales = {h.dia_semana: h for h in clase.horarios.all()}
    return render_template('editar_horarios.html', clase=clase, horarios_actuales=horarios_actuales)

# ---------------------
# RUTA TEMPORAL PARA CREAR PRIMER ADMIN
# ---------------------
# IMPORTANTE: Elimina o protege esta ruta después de crear el primer admin
@app.route('/crear-primer-admin', methods=['GET', 'POST'])
def crear_primer_admin():
    """Ruta temporal para crear el primer usuario administrador"""
    # Verificar si ya existe un admin
    admin_existente = Usuario.query.filter_by(rol='admin').first()
    if admin_existente:
        return '''
        <html>
        <body style="font-family: Arial; padding: 50px; text-align: center;">
            <h2>Ya existe un administrador</h2>
            <p>El usuario administrador ya ha sido creado.</p>
            <p>Usuario: <strong>{}</strong></p>
            <p><a href="/login">Ir al login</a></p>
            <hr>
            <p style="color: red; font-size: 12px;">
                Por seguridad, elimina o protege esta ruta después de crear el primer admin.
            </p>
        </body>
        </html>
        '''.format(admin_existente.username)
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        password_confirm = request.form.get('password_confirm', '').strip()
        nombre_completo = request.form.get('nombre_completo', '').strip()
        
        errores = []
        
        if not username:
            errores.append('El nombre de usuario es obligatorio')
        elif Usuario.query.filter_by(username=username).first():
            errores.append('El nombre de usuario ya existe')
        
        if not email:
            errores.append('El correo electrónico es obligatorio')
        elif Usuario.query.filter_by(email=email).first():
            errores.append('El correo electrónico ya está registrado')
        
        if not password:
            errores.append('La contraseña es obligatoria')
        elif len(password) < 6:
            errores.append('La contraseña debe tener al menos 6 caracteres')
        
        if password != password_confirm:
            errores.append('Las contraseñas no coinciden')
        
        if not nombre_completo:
            errores.append('El nombre completo es obligatorio')
        
        if errores:
            return render_template('crear_admin_temp.html', errores=errores, 
                                 username=username, email=email, nombre_completo=nombre_completo)
        
        # Crear usuario admin
        nuevo_admin = Usuario(
            username=username,
            email=email,
            nombre_completo=nombre_completo,
            rol='admin'
        )
        nuevo_admin.set_password(password)
        
        try:
            db.session.add(nuevo_admin)
            db.session.commit()
            return '''
            <html>
            <body style="font-family: Arial; padding: 50px; text-align: center;">
                <h2 style="color: green;">✓ Administrador creado exitosamente</h2>
                <p><strong>Usuario:</strong> {}</p>
                <p><strong>Email:</strong> {}</p>
                <p><strong>Nombre:</strong> {}</p>
                <hr>
                <p><a href="/login" style="background: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Ir al Login</a></p>
                <hr>
                <p style="color: red; font-size: 12px;">
                    ⚠️ IMPORTANTE: Por seguridad, elimina o protege la ruta /crear-primer-admin después de crear el primer admin.
                </p>
            </body>
            </html>
            '''.format(username, email, nombre_completo)
        except Exception as e:
            db.session.rollback()
            return f'<h2>Error al crear administrador: {str(e)}</h2><a href="/crear-primer-admin">Intentar de nuevo</a>'
    
    return render_template('crear_admin_temp.html', errores=[], username='', email='', nombre_completo='')

# Inicializar DB (solo si quieres crear tablas desde aquí)
@app.cli.command("init-db")
def init_db():
    db.create_all()
    print("Tablas creadas")

if __name__ == '__main__':
    # Verificar si existen certificados SSL para HTTPS
    cert_dir = os.path.join(os.path.dirname(__file__), 'certificados')
    cert_file = os.path.join(cert_dir, 'cert.pem')
    key_file = os.path.join(cert_dir, 'key.pem')
    
    use_https = os.path.exists(cert_file) and os.path.exists(key_file)
    
    # Evitar ejecución duplicada cuando Flask usa el reloader
    # Solo mostrar mensajes e inicializar scheduler en el proceso hijo
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        if use_https:
            # Obtener IP local para mostrar en el mensaje
            import socket
            hostname = socket.gethostname()
            try:
                local_ip = socket.gethostbyname(hostname)
            except:
                local_ip = "192.168.100.34"
            
            print("=" * 60)
            print("Iniciando servidor Flask con HTTPS")
            print("=" * 60)
            print(f"\nURLs de acceso:")
            print(f"  Desde tu computadora: https://localhost:5000")
            print(f"  Desde iPhone/otros dispositivos: https://{local_ip}:5000")
            print("=" * 60)
            print("\n[!] IMPORTANTE para iPhone:")
            print("   1. Asegurate de que iPhone y computadora esten en la misma red WiFi")
            print("   2. Usa Safari (no Chrome) en iPhone")
            print("   3. Escribe: https://" + local_ip + ":5000")
            print("   4. Safari mostrara advertencia - toca 'Avanzado' > 'Continuar'")
            print("=" * 60)
            print("\nServidor iniciando...")
            print("Presiona CTRL+C para detener")
            print("=" * 60)
        else:
            print("=" * 60)
            print("⚠️  Iniciando servidor Flask sin HTTPS")
            print("=" * 60)
            print("📱 Para usar en iPhone, necesitas HTTPS.")
            print("   Ejecuta: python generar_certificados.py")
            print("   Luego reinicia el servidor.")
            print("=" * 60)
        
        # Inicializar el scheduler antes de iniciar el servidor (solo en proceso hijo)
        if scheduler is None:
            scheduler = BackgroundScheduler()
            scheduler.add_job(
                func=verificar_y_enviar_reportes,
                trigger=CronTrigger(minute='*'),  # Ejecutar cada minuto
                id='verificar_reportes',
                name='Verificar y enviar reportes de asistencia',
                replace_existing=True
            )
            scheduler.start()
            atexit.register(lambda: scheduler.shutdown() if scheduler else None)
            print("✅ Scheduler de reportes automáticos iniciado")
    
    # Ejecutar el servidor siempre (tanto en proceso padre como hijo)
    if use_https:
        try:
            app.run(
                host='0.0.0.0', 
                port=int(os.getenv('PORT', 5000)), 
                debug=True,
                ssl_context=(cert_file, key_file)
            )
        except OSError as e:
            if "Address already in use" in str(e) or "address is already in use" in str(e).lower():
                print("\n[!] ERROR: El puerto 5000 ya esta en uso.")
                print("   Cierra la otra aplicacion que esta usando el puerto 5000")
                print("   O cambia el puerto en la variable de entorno PORT")
            else:
                print(f"\n[!] ERROR al iniciar el servidor: {e}")
                raise
    else:
        app.run(host='0.0.0.0', port=int(os.getenv('PORT', 5000)), debug=True)
